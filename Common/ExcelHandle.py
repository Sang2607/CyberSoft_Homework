from openpyxl import Workbook,load_workbook
import os
from Common.FunctionHandle import inputStr
from Common.Decorator import handle_exception
#insert
def CreateExcel(FileName,Title,NTL):
    wb = Workbook()
    ws = wb.active
    ws.title = Title
    if len(NTL) > 0:
        for i in NTL:
            ws.append(i)
    wb.save(FileName)

#Append data
def AppendRows(FileName,Title,NTL):
    wb = load_workbook(FileName)
    ws = wb[Title]
    for i in NTL:
        ws.append(i)
    wb.save(FileName)

@handle_exception
def SaveFile(FileName,Title,NestedList,dialogConfirm:str = None,removeHeder = True):
    if dialogConfirm and inputStr(dialogConfirm + " (y/n):") != "y":
        return
    if os.path.exists(FileName):
        if len(NestedList) == 0:
            print("Lỗi: Vui lòng kiểm tra Data cần thêm")
        else:
            AppendRows(FileName,Title, NestedList[1:] if removeHeder else NestedList)
    else:
        CreateExcel(FileName,Title,NestedList)
    print("Thành công!")

#đọc tất cả dữ liệu trong file
def ReadFile(FileName):
    print(FileName)
    wb = load_workbook(FileName)
    ws = wb.active

    for row in ws.iter_rows(values_only=True):
        print(row)
